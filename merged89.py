import csv
import json
import os
import queue
import sqlite3
import threading
import time
from datetime import datetime, time as dtime, timedelta
from typing import Any, Dict, List, Optional

import cv2
import numpy as np
from flask import Flask, Response, jsonify, render_template_string, request, send_file, stream_with_context

try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except ImportError:
    pass  # python-dotenv is optional; env vars can also be set directly

# ------------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------------

# Camera connection details come from environment variables (see .env.example)
# instead of being hardcoded here, so real credentials/IPs never end up in
# git history. This is "Hotsaw 1" - run it with its own .env pointing at
# the .89 camera (merged88.py is "Hotsaw 2", pointing at .88).
CAM_USER = os.environ.get("HOTSAW_CAM_USER", "admin")
CAM_PASS = os.environ.get("HOTSAW_CAM_PASS", "changeme")
CAM_HOST = os.environ.get("HOTSAW_CAM_HOST", "127.0.0.1")
CAM_PORT = os.environ.get("HOTSAW_CAM_PORT", "554")

RTSP_URL = f"rtsp://{CAM_USER}:{CAM_PASS}@{CAM_HOST}:{CAM_PORT}/cam/realmonitor?channel=1&subtype=0"

GST_STR = (
    f"rtspsrc location=rtsp://{CAM_USER}:{CAM_PASS}@{CAM_HOST}:{CAM_PORT}/cam/realmonitor?channel=1&subtype=0 "
    r"latency=100 ! queue ! rtph264depay ! h264parse ! avdec_h264 ! videoconvert ! videoscale ! "
    r"video/x-raw,width=1280,height=720 ! appsink"
)

SQLITE_DB = "shared_data.db"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
YOLO_ENDFACE   = os.path.join(BASE_DIR, "endfacemv2.pt")
YOLO_FRONTFACE = os.path.join(BASE_DIR, "frontfacedetect1.pt")


ITEMS = [
    "MB 150", "MB 200", "MB 250", "MB 300", "MB 350", "MB 400",
    "MC 150", "MC 175", "MC 200", "MC 250", "MC 300",
    "Angle 130*10/12", "Angle 150*10/12",
]

for _d in ("cut_events", "debug_frames", "motion_masks", "bright_masks", "logs", "reports"):
    os.makedirs(_d, exist_ok=True)

REPORT_DIR = "reports_hotsaw1"
CURRENT_STATE_PATH = os.path.join(REPORT_DIR, "current_running_session.json")

# ── Completed-session reports (appended on reset) ──────────────────
BLADE_CSV_PATH  = os.path.join(REPORT_DIR, "blade_reset_report.csv")
BLADE_XLSX_PATH = os.path.join(REPORT_DIR, "blade_reset_report.xlsx")
ITEM_CSV_PATH   = os.path.join(REPORT_DIR, "item_wise_reset_report.csv")
ITEM_XLSX_PATH  = os.path.join(REPORT_DIR, "item_wise_reset_report.xlsx")

# ── Live-session files (overwritten after every cut) ───────────────
BLADE_LIVE_CSV_PATH = os.path.join(REPORT_DIR, "blade_live_session.csv")
ITEM_LIVE_CSV_PATH  = os.path.join(REPORT_DIR, "item_live_session.csv")

# Blade Resets sheet — one row per blade reset (Report File column removed)
BLADE_HEADERS = [
    "ID", "Date", "Time", "Blade Count", "Blade Start", "Blade End",
    "Current Item", "Item Breakdown",
]

# Item Wise sheet — one row per blade SESSION, one column per item.
# The row is created when a blade session starts and is updated in place
# on every item reset / item change (never appended to per-reset).
ITEM_HEADERS = [
    "ID", "Date", "Time",
    "MB 150", "MB 200", "MB 250", "MB 300", "MB 350", "MB 400",
    "MC 150", "MC 175", "MC 200", "MC 250", "MC 300",
    "Angle 130*10/12", "Angle 150*10/12",
]

COMBINED_XLSX_PATH = os.path.join(REPORT_DIR, "msp_hotsow_combined_report.xlsx")
BLADE_LIVE_HEADERS = ["ID", "Session_Start", "Last_Updated", "Blade_Count", "Current_Item", "Item_Breakdown"]
ITEM_LIVE_HEADERS  = ["ID", "Item", "Session_Start", "Last_Updated", "Item_Count", "Blade_Count"]

# ------------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------------

def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _date_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")

def _time_str() -> str:
    return datetime.now().strftime("%H:%M:%S")

def get_previous_date() -> str:
    return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

def _safe_int(v: Any, default: int = 0) -> int:
    try:
        return max(0, int(v or 0))
    except Exception:
        return default

def _default_runtime_state() -> Dict[str, Any]:
    now = _now_str()
    return {
        "blade_count": 0,
        "blade_start_time": now,
        "current_item": ITEMS[0],
        "item_session_start_time": now,
        "item_counts": {item: 0 for item in ITEMS},
        "blade_item_counts": {item: 0 for item in ITEMS},
    }

def load_runtime_state() -> Dict[str, Any]:
    default = _default_runtime_state()
    try:
        if not os.path.isfile(CURRENT_STATE_PATH):
            return default
        with open(CURRENT_STATE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        blade_count   = _safe_int(data.get("blade_count", data.get("cut_count", 0)))
        blade_start   = str(data.get("blade_start_time", data.get("session_start_time", default["blade_start_time"])))
        current_item  = str(data.get("current_item", default["current_item"]))
        if current_item not in ITEMS:
            current_item = default["current_item"]
        item_start         = str(data.get("item_session_start_time", default["item_session_start_time"]))
        item_counts        = {item: _safe_int((data.get("item_counts") or {}).get(item, 0)) for item in ITEMS}
        blade_item_counts  = {item: _safe_int((data.get("blade_item_counts") or {}).get(item, 0)) for item in ITEMS}
        return {
            "blade_count": blade_count,
            "blade_start_time": blade_start,
            "current_item": current_item,
            "item_session_start_time": item_start,
            "item_counts": item_counts,
            "blade_item_counts": blade_item_counts,
        }
    except Exception as e:
        print(f"[WARN] Could not load runtime state: {e}")
        return default

def save_runtime_state() -> None:
    try:
        os.makedirs(REPORT_DIR, exist_ok=True)
        with _state_lock:
            data = {
                "blade_count": int(_state.get("blade_count", 0)),
                "cut_count":   int(_state.get("blade_count", 0)),
                "blade_start_time":       _state.get("blade_start_time", _now_str()),
                "session_start_time":     _state.get("blade_start_time", _now_str()),
                "current_item":           _state.get("current_item", ITEMS[0]),
                "item_session_start_time":_state.get("item_session_start_time", _now_str()),
                "item_counts":            dict(_state.get("item_counts", {})),
                "blade_item_counts":      dict(_state.get("blade_item_counts", {})),
                "updated_at":             _now_str(),
            }
        tmp = CURRENT_STATE_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        os.replace(tmp, CURRENT_STATE_PATH)
    except Exception as e:
        print(f"[ERROR] save_runtime_state: {e}")


# ── Live CSV writers (overwritten after every cut) ─────────────────

def write_live_blade_state() -> None:
    """Overwrite blade_live_session.csv with current session state after every cut."""
    try:
        with _state_lock:
            blade_count       = int(_state.get("blade_count", 0))
            blade_start       = str(_state.get("blade_start_time", _now_str()))
            current_item      = str(_state.get("current_item", ITEMS[0]))
            blade_item_counts = dict(_state.get("blade_item_counts", {}))
        breakdown = json.dumps(
            {k: v for k, v in blade_item_counts.items() if int(v or 0) > 0},
            ensure_ascii=False,
        )
        with open(BLADE_LIVE_CSV_PATH, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(BLADE_LIVE_HEADERS)
            w.writerow(["LIVE", blade_start, _now_str(), blade_count, current_item, breakdown])
    except Exception as e:
        print(f"[WARN] write_live_blade_state: {e}")


def write_live_item_state() -> None:
    """Overwrite item_live_session.csv with current item session state after every cut."""
    try:
        with _state_lock:
            current_item = str(_state.get("current_item", ITEMS[0]))
            item_start   = str(_state.get("item_session_start_time", _now_str()))
            item_count   = int(_state.get("item_counts", {}).get(current_item, 0))
            blade_count  = int(_state.get("blade_count", 0))
        with open(ITEM_LIVE_CSV_PATH, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(ITEM_LIVE_HEADERS)
            w.writerow(["LIVE", current_item, item_start, _now_str(), item_count, blade_count])
    except Exception as e:
        print(f"[WARN] write_live_item_state: {e}")


# ── Report CSV/XLSX helpers ────────────────────────────────────────

def ensure_csv_header(path: str, headers: List[str]) -> None:
    """Make sure `path` exists with exactly `headers` as its header row.

    If the file already exists but was written by an OLDER schema (e.g. the
    previous "Item / Item Count / Start / End" layout instead of the new
    per-item-column layout), the mismatched file is archived (renamed with
    a timestamp suffix) and a fresh file with the correct header is
    created. Without this, an old file silently keeps its old columns
    forever and the dashboard/Excel show blank/incorrect data because
    they're reading columns that no longer exist under those names.
    """
    os.makedirs(REPORT_DIR, exist_ok=True)
    if not os.path.isfile(path) or os.path.getsize(path) == 0:
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(headers)
        return
    try:
        with open(path, "r", newline="", encoding="utf-8") as f:
            first_row = next(csv.reader(f), None)
    except Exception as e:
        print(f"[WARN] ensure_csv_header could not read {path}: {e}")
        first_row = None
    if first_row != headers:
        backup_path = f"{path}.old_{int(time.time())}.bak"
        try:
            os.replace(path, backup_path)
            print(f"[MIGRATE] {path} had an outdated schema. Old file backed up to {backup_path}.")
        except Exception as e:
            print(f"[ERROR] Failed to back up outdated {path}: {e}")
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(headers)

def read_csv_logs(path: str, headers: List[str]) -> List[Dict[str, Any]]:
    ensure_csv_header(path, headers)
    rows: List[Dict[str, Any]] = []
    try:
        with open(path, "r", newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                # A row with MORE fields than the header (two rows written
                # to the same physical CSV line without a newline between
                # them) makes DictReader stash the extras under a literal
                # None key, which then crashes Flask's JSON encoder
                # (sort_keys=True can't compare None to a str) and 500s the
                # WHOLE /api/blade_logs or /api/item_logs response, not
                # just the one bad row. Drop it and log a warning instead.
                if None in row:
                    print(f"[WARN] {path}: row ID={row.get('ID')} has more "
                          f"columns than the header ({row.pop(None)!r} "
                          f"ignored) - file may have a corrupted/merged line.")
                rows.append(dict(row))
    except Exception as e:
        print(f"[ERROR] read_csv_logs {path}: {e}")
    rows.sort(key=lambda r: _safe_int(r.get("ID", 0)), reverse=True)
    return rows


def next_id(path: str, headers: List[str]) -> int:
    rows = read_csv_logs(path, headers)
    return 1 if not rows else max(_safe_int(r.get("ID", 0)) for r in rows) + 1

def sync_excel(csv_path: str, xlsx_path: str, sheet_name: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        with open(csv_path, "r", newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                ws.append(row)
        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)
        wb.save(xlsx_path)
    except Exception as e:
        print(f"[WARN] Excel sync failed for {csv_path}. CSV still saved. Reason: {e}")

def build_combined_excel() -> str:
    """Single workbook: Sheet1 'Blade Resets' + Sheet2 'Item Wise', linked by ID <-> Blade_ID."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    ensure_csv_header(BLADE_CSV_PATH, BLADE_HEADERS)
    ensure_csv_header(ITEM_CSV_PATH, ITEM_HEADERS)
    wb = Workbook()

    def _fill_sheet(ws, csv_path, title):
        ws.title = title[:31]
        with open(csv_path, "r", newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                ws.append(row)
        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)

    _fill_sheet(wb.active, BLADE_CSV_PATH, "Blade Resets")
    _fill_sheet(wb.create_sheet(), ITEM_CSV_PATH, "Item Wise")
    wb.save(COMBINED_XLSX_PATH)
    return COMBINED_XLSX_PATH

def append_blade_report(blade_count: int, blade_start: str, current_item: str, item_breakdown: Dict[str, int]) -> Dict[str, Any]:
    breakdown = json.dumps({k: v for k, v in item_breakdown.items() if int(v) > 0}, ensure_ascii=False)
    with _csv_lock:
        ensure_csv_header(BLADE_CSV_PATH, BLADE_HEADERS)
        rid = next_id(BLADE_CSV_PATH, BLADE_HEADERS)
        row = [rid, _date_str(), _time_str(), blade_count, blade_start, _now_str(), current_item, breakdown]
        with open(BLADE_CSV_PATH, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(row)
        sync_excel(BLADE_CSV_PATH, BLADE_XLSX_PATH, "Blade Reset Report")
    return {
        "id": rid, "date": row[1], "time": row[2], "blade_count": blade_count,
        "blade_start": blade_start, "blade_end": row[5], "current_item": current_item,
        "item_breakdown": breakdown,
    }


# ── Item Wise session-row helpers ───────────────────────────────────
# The Item Wise sheet keeps ONE ROW PER BLADE SESSION, with one column
# per item. A new row is created only when a blade session starts
# (i.e. right after a blade reset). Item resets and item changes update
# that same row in place instead of appending new rows.

def create_new_item_session_row() -> Dict[str, Any]:
    """Append a fresh, all-zero row to the Item Wise CSV for a new blade session."""
    with _csv_lock:
        ensure_csv_header(ITEM_CSV_PATH, ITEM_HEADERS)
        rid = next_id(ITEM_CSV_PATH, ITEM_HEADERS)
        row = [rid, _date_str(), _time_str()] + [0 for _ in ITEMS]
        with open(ITEM_CSV_PATH, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(row)
        sync_excel(ITEM_CSV_PATH, ITEM_XLSX_PATH, "Item Wise Report")
    return {"id": rid, "date": row[1], "time": row[2]}

def update_item_session(item_name: str, count: int) -> None:
    """Update the LAST ROW of the Item Wise CSV with the given item's count."""
    if item_name not in ITEMS:
        return
    with _csv_lock:
        ensure_csv_header(ITEM_CSV_PATH, ITEM_HEADERS)
        try:
            with open(ITEM_CSV_PATH, "r", newline="", encoding="utf-8") as f:
                rows = list(csv.reader(f))
        except Exception as e:
            print(f"[ERROR] update_item_session read: {e}")
            return

        if len(rows) < 2:
            # No session row exists yet — create one first, then re-read.
            # (create_new_item_session_row also takes _csv_lock, so its
            # body is inlined here to avoid deadlocking on re-entry.)
            rid = next_id(ITEM_CSV_PATH, ITEM_HEADERS)
            new_row = [rid, _date_str(), _time_str()] + [0 for _ in ITEMS]
            with open(ITEM_CSV_PATH, "a", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(new_row)
            with open(ITEM_CSV_PATH, "r", newline="", encoding="utf-8") as f:
                rows = list(csv.reader(f))

        header = rows[0]
        last_row = rows[-1]
        try:
            col_idx = header.index(item_name)
        except ValueError:
            return

        while len(last_row) < len(header):
            last_row.append(0)

        last_row[col_idx] = count
        last_row[2] = _time_str()  # keep "Time" fresh on every update
        rows[-1] = last_row

        with open(ITEM_CSV_PATH, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)
    sync_excel(ITEM_CSV_PATH, ITEM_XLSX_PATH, "Item Wise Report")

# ------------------------------------------------------------------
# CUT COUNTER THRESHOLDS
# ------------------------------------------------------------------
ROI_X1, ROI_Y1 = 170, 90
ROI_X2, ROI_Y2 = 560, 350
BLADE_X1, BLADE_Y1 = 120, 105
BLADE_X2, BLADE_Y2 = 190, 210

BRIGHT_THRESHOLD  = 10_000
MOTION_THRESHOLD  = 18_000
SHARP_THRESHOLD   = 350
EXTREME_BRIGHT    = 25_000
EXTREME_MOTION    = 50_000
TRUE_CUT_OVERLAP  = 700
TRUE_CUT_SPARK    = 450
TRUE_CUT_MOTION   = 1_800
MIN_CUT_FRAMES    = 4
RELEASE_FRAMES    = 5
CUT_COOLDOWN      = 3.0

# ------------------------------------------------------------------
# SHARED STATE
# ------------------------------------------------------------------
_state_lock = threading.Lock()
_csv_lock = threading.Lock()  # guards every write to the report CSVs so two
                               # near-simultaneous blade/item events can't
                               # interleave their writes onto the same
                               # physical CSV line (that corruption caused a
                               # 500 error on /api/blade_logs previously).
_runtime = load_runtime_state()
_state: Dict[str, Any] = {
    "blade_count":            _runtime["blade_count"],
    "cut_count":              _runtime["blade_count"],
    "bright": 0, "motion": 0, "sharpness": 0, "score": 0,
    "cut_detected": False, "frame_id": 0,
    "blade_start_time":        _runtime["blade_start_time"],
    "session_start_time":      _runtime["blade_start_time"],
    "current_item":            _runtime["current_item"],
    "item_session_start_time": _runtime["item_session_start_time"],
    "item_counts":             _runtime["item_counts"],
    "blade_item_counts":       _runtime["blade_item_counts"],
}

_billet_lock = threading.Lock()
_billet_state: Dict[str, Any] = {
    "billet_len": 0, "final_len": 0, "billet_count": 0,
    "len_6": 0, "len_9": 0, "len_12": 0, "uncat": 0,
}

_sse_clients: List[queue.Queue] = []
_sse_lock = threading.Lock()
_latest_frame_lock = threading.Lock()
_latest_frame: Optional[bytes] = None
_latest_gst_frame_lock = threading.Lock()
_latest_gst_frame: Optional[bytes] = None

# Blade reset flag — False or str (next_item after reset)
_reset_blade_flag_lock = threading.Lock()
_reset_blade_flag: Any = False
_reset_blade_next_item: str = ""

# Change item (no reset) flag
_change_item_flag_lock = threading.Lock()
_change_item_flag: bool = False
_change_item_new_item: str = ""

save_runtime_state()

# Make sure the Item Wise sheet always has an active session row to update.
ensure_csv_header(ITEM_CSV_PATH, ITEM_HEADERS)
if not read_csv_logs(ITEM_CSV_PATH, ITEM_HEADERS):
    create_new_item_session_row()

# ------------------------------------------------------------------
# SSE
# ------------------------------------------------------------------
def _broadcast_event(event_type: str, data: str) -> None:
    msg = f"event: {event_type}\ndata: {data}\n\n"
    with _sse_lock:
        dead = []
        for q in _sse_clients:
            try:
                q.put_nowait(msg)
            except queue.Full:
                dead.append(q)
        for q in dead:
            if q in _sse_clients:
                _sse_clients.remove(q)

def _state_payload() -> Dict[str, Any]:
    with _state_lock:
        item        = str(_state.get("current_item", ITEMS[0]))
        item_counts = dict(_state.get("item_counts", {}))
        return {
            "blade_count":             int(_state.get("blade_count", 0)),
            "cut_count":               int(_state.get("blade_count", 0)),
            "bright":                  int(_state.get("bright", 0)),
            "motion":                  int(_state.get("motion", 0)),
            "sharpness":               int(_state.get("sharpness", 0)),
            "score":                   int(_state.get("score", 0)),
            "frame_id":                int(_state.get("frame_id", 0)),
            "blade_start_time":        _state.get("blade_start_time", ""),
            "session_start_time":      _state.get("blade_start_time", ""),
            "current_item":            item,
            "current_item_count":      int(item_counts.get(item, 0)),
            "item_session_start_time": _state.get("item_session_start_time", ""),
            "item_counts":             item_counts,
            "blade_item_counts":       dict(_state.get("blade_item_counts", {})),
        }

# ------------------------------------------------------------------
# VIDEO OVERLAY  –  new format:
#   BLADE COUNT: {BC}   {ITEM} --> {ITEM COUNT}
# ------------------------------------------------------------------
def draw_video_overlay(frame: np.ndarray, payload: Dict[str, Any]) -> np.ndarray:
    blade_count  = int(payload.get("blade_count", 0) or 0)
    current_item = str(payload.get("current_item", "-"))
    item_count   = int(payload.get("current_item_count", 0) or 0)

    h, w = frame.shape[:2]

    # semi-transparent background bar
    overlay = frame.copy()
    cv2.rectangle(overlay, (0, 0), (w, 60), (0, 0, 0), -1)
    frame = cv2.addWeighted(overlay, 0.55, frame, 0.45, 0)

    # ── left block: BLADE COUNT: {BC} ──
    bc_label = "BLADE COUNT:"
    bc_value = str(blade_count)
    cv2.putText(frame, bc_label, (14, 42),
                cv2.FONT_HERSHEY_SIMPLEX, 0.95, (180, 180, 180), 2, cv2.LINE_AA)
    # measure label width so value sits right after it
    (lw, _), _ = cv2.getTextSize(bc_label, cv2.FONT_HERSHEY_SIMPLEX, 0.95, 2)
    cv2.putText(frame, bc_value, (14 + lw + 8, 42),
                cv2.FONT_HERSHEY_SIMPLEX, 1.05, (0, 165, 255), 3, cv2.LINE_AA)

    # ── right block: {ITEM} --> {ITEM COUNT} ──
    arrow_str   = f"{current_item}  -->  {item_count}"
    (aw, _), _  = cv2.getTextSize(arrow_str, cv2.FONT_HERSHEY_SIMPLEX, 0.95, 2)
    x_right     = w - aw - 14

    # draw item name in white
    item_str  = f"{current_item}  --> "
    (iw, _), _ = cv2.getTextSize(item_str, cv2.FONT_HERSHEY_SIMPLEX, 0.95, 2)
    cv2.putText(frame, item_str, (x_right, 42),
                cv2.FONT_HERSHEY_SIMPLEX, 0.95, (255, 255, 255), 2, cv2.LINE_AA)
    # draw count in green
    cv2.putText(frame, str(item_count), (x_right + iw, 42),
                cv2.FONT_HERSHEY_SIMPLEX, 1.05, (0, 255, 128), 3, cv2.LINE_AA)

    return frame

# ------------------------------------------------------------------
# THREAD 1 — CUT COUNTER DETECTION
# ------------------------------------------------------------------
def detection_loop() -> None:
    global _latest_frame, _reset_blade_flag, _reset_blade_next_item
    global _change_item_flag, _change_item_new_item

    cap = cv2.VideoCapture(RTSP_URL)
    prev_gray = None
    with _state_lock:
        blade_count = int(_state.get("blade_count", 0))
    last_cut = time.time()
    frame_id = 0
    cut_active = False
    cut_signal_frames = 0
    no_cut_frames = 0
    spark_history: List[int] = []
    motion_history: List[int] = []
    kernel = np.ones((3, 3), np.uint8)

    while True:
        # ── Handle blade reset request ─────────────────────────────
        with _reset_blade_flag_lock:
            if _reset_blade_flag:
                next_item_after_reset = _reset_blade_next_item
                blade_count = 0
                cut_active  = False
                new_start   = _now_str()
                with _state_lock:
                    _state["blade_count"]       = 0
                    _state["cut_count"]         = 0
                    _state["blade_start_time"]  = new_start
                    _state["session_start_time"]= new_start
                    _state["blade_item_counts"] = {item: 0 for item in ITEMS}
                    _state["item_counts"]       = {item: 0 for item in ITEMS}
                    if next_item_after_reset in ITEMS:
                        _state["current_item"]            = next_item_after_reset
                        _state["item_session_start_time"] = new_start
                save_runtime_state()
                write_live_blade_state()
                write_live_item_state()
                _reset_blade_flag      = False
                _reset_blade_next_item = ""
                _broadcast_event("blade_reset", json.dumps(_state_payload()))

        # ── Handle change-item-only (no blade reset) ───────────────
        with _change_item_flag_lock:
            if _change_item_flag:
                new_item = _change_item_new_item
                if new_item in ITEMS:
                    with _state_lock:
                        _state["current_item"]            = new_item
                        _state["item_session_start_time"] = _now_str()
                        _state["item_counts"].setdefault(new_item, 0)
                        _state["blade_item_counts"].setdefault(new_item, 0)
                    save_runtime_state()
                    write_live_item_state()
                    _broadcast_event("item_changed", json.dumps(_state_payload()))
                    print(f"[ITEM CHANGE] Switched to: {new_item} (blade count unchanged)")
                _change_item_flag     = False
                _change_item_new_item = ""

        ret, frame = cap.read()
        if not ret:
            print("[CUT] Camera disconnect. Reconnecting...")
            try:
                cap.release()
            except Exception:
                pass
            time.sleep(2)
            cap = cv2.VideoCapture(RTSP_URL)
            prev_gray = None
            continue

        frame_id += 1
        frame = cv2.resize(frame, (1280, 720))
        roi   = frame[ROI_Y1:ROI_Y2, ROI_X1:ROI_X2]
        gray  = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        blur  = cv2.GaussianBlur(gray, (5, 5), 0)

        _, bright_mask = cv2.threshold(blur, 210, 255, cv2.THRESH_BINARY)
        bright_mask    = cv2.morphologyEx(bright_mask, cv2.MORPH_OPEN, kernel)
        bright_pixels  = cv2.countNonZero(bright_mask)
        sharpness      = cv2.Laplacian(gray, cv2.CV_64F).var()

        motion_pixels = 0
        motion_mask   = np.zeros_like(gray)
        if prev_gray is not None:
            diff = cv2.absdiff(prev_gray, gray)
            _, motion_mask = cv2.threshold(diff, 25, 255, cv2.THRESH_BINARY)
            motion_mask    = cv2.dilate(motion_mask, kernel, iterations=2)
            motion_pixels  = cv2.countNonZero(motion_mask)
        prev_gray = gray.copy()

        spark_history.append(bright_pixels)
        motion_history.append(motion_pixels)
        if len(spark_history)  > 5: spark_history.pop(0)
        if len(motion_history) > 5: motion_history.pop(0)
        avg_bright = int(np.mean(spark_history))
        avg_motion = int(np.mean(motion_history))

        score = 0
        if avg_bright   > BRIGHT_THRESHOLD:  score += 1
        if avg_motion   > MOTION_THRESHOLD:  score += 1
        if sharpness    > SHARP_THRESHOLD:   score += 1
        if bright_pixels > EXTREME_BRIGHT:   score += 2
        if motion_pixels > EXTREME_MOTION:   score += 2

        hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
        sm1 = cv2.inRange(hsv, (5, 80, 150), (45, 255, 255))
        sm2 = cv2.inRange(hsv, (0, 110, 180), (8, 255, 255))
        spark_mask = cv2.bitwise_or(sm1, sm2)
        spark_mask = cv2.morphologyEx(spark_mask, cv2.MORPH_OPEN, kernel)
        spark_mask = cv2.dilate(spark_mask, kernel, iterations=1)

        blade_spark  = spark_mask[BLADE_Y1:BLADE_Y2, BLADE_X1:BLADE_X2]
        blade_motion = motion_mask[BLADE_Y1:BLADE_Y2, BLADE_X1:BLADE_X2]
        blade_overlap = cv2.bitwise_and(blade_spark, blade_motion)

        blade_spark_px  = cv2.countNonZero(blade_spark)
        blade_motion_px = cv2.countNonZero(blade_motion)
        overlap_px      = cv2.countNonZero(blade_overlap)

        raw_cut_signal = (
            blade_spark_px  > TRUE_CUT_SPARK  and
            blade_motion_px > TRUE_CUT_MOTION and
            overlap_px      > TRUE_CUT_OVERLAP
        )

        if raw_cut_signal:
            cut_signal_frames += 1
            no_cut_frames = 0
        else:
            no_cut_frames += 1
            if no_cut_frames >= RELEASE_FRAMES:
                cut_signal_frames = 0

        true_cut_signal = cut_signal_frames >= MIN_CUT_FRAMES
        now = time.time()
        cut_detected = False

        if true_cut_signal and not cut_active and (now - last_cut > CUT_COOLDOWN):
            blade_count += 1
            with _state_lock:
                current_item = str(_state.get("current_item", ITEMS[0]))
                _state["blade_count"] = blade_count
                _state["cut_count"]   = blade_count
                _state["item_counts"][current_item]       = int(_state["item_counts"].get(current_item, 0)) + 1
                _state["blade_item_counts"][current_item] = int(_state["blade_item_counts"].get(current_item, 0)) + 1
                current_item_count = int(_state["item_counts"][current_item])
            save_runtime_state()
            write_live_blade_state()
            write_live_item_state()
            try:
                update_item_session(current_item, current_item_count)
            except Exception as e:
                print(f"[WARN] update_item_session on cut failed: {e}")

            last_cut     = now
            cut_active   = True
            cut_detected = True
            cv2.imwrite(f"cut_events/cut_{blade_count}.jpg",     frame)
            cv2.imwrite(f"debug_frames/debug_{blade_count}.jpg",  roi.copy())
            cv2.imwrite(f"motion_masks/motion_{blade_count}.jpg", motion_mask)
            cv2.imwrite(f"bright_masks/bright_{blade_count}.jpg", spark_mask)
            print(f"[CUT #{blade_count}] Item:{current_item} Spark:{blade_spark_px} Motion:{blade_motion_px} Overlap:{overlap_px}")
            _broadcast_event("cut", json.dumps(_state_payload()))

        if no_cut_frames >= RELEASE_FRAMES:
            cut_active = False

        with _state_lock:
            _state.update({
                "blade_count":  blade_count,
                "cut_count":    blade_count,
                "bright":       bright_pixels,
                "motion":       motion_pixels,
                "sharpness":    int(sharpness),
                "score":        score,
                "cut_detected": cut_detected,
                "frame_id":     frame_id,
            })

        if frame_id % 3 == 0:
            _broadcast_event("metrics", json.dumps(_state_payload()))

        frame = draw_video_overlay(frame, _state_payload())
        ok, jpeg = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
        if ok:
            with _latest_frame_lock:
                _latest_frame = jpeg.tobytes()

# ------------------------------------------------------------------
# THREAD 2 — GSTREAMER / YOLO DETECTION
# ------------------------------------------------------------------
def gstreamer_loop() -> None:
    global _latest_gst_frame

    yolo_available = False
    yolo_error = ""
    model_endface = model_frontface = None
    cls_endface, cls_frontface = ["endface"], ["frontface"]
    YOLO = None

    # NOTE: the old code required BOTH weight files to be present/loadable or
    # it disabled YOLO entirely - so if just one of the two files was
    # missing/renamed/corrupt, or ultralytics failed to load one of them,
    # you'd get "YOLO OFF" with NO detection at all, even though the other
    # model was perfectly fine. Each model is now loaded independently so a
    # problem with one doesn't take down the other, and the real per-model
    # error shows up in the console/status overlay instead of a generic
    # "file(s) not found" message that doesn't say which one.
    print(f"[GST] Looking for YOLO weights at: {YOLO_ENDFACE} and {YOLO_FRONTFACE}")
    try:
        from ultralytics import YOLO  # type: ignore
    except Exception as e:
        yolo_error = f"ultralytics import failed: {type(e).__name__}: {e}"
        print(f"[GST] YOLO unavailable ({yolo_error}). Streaming raw frames.")

    if YOLO is not None:
        errors = []
        if os.path.isfile(YOLO_ENDFACE):
            try:
                model_endface = YOLO(YOLO_ENDFACE)
                print(f"[GST] endface model loaded. classes={model_endface.names}")
            except Exception as e:
                errors.append(f"endface: {type(e).__name__}: {e}")
        else:
            errors.append(f"endface: weights file not found at {YOLO_ENDFACE}")

        if os.path.isfile(YOLO_FRONTFACE):
            try:
                model_frontface = YOLO(YOLO_FRONTFACE)
                print(f"[GST] frontface model loaded. classes={model_frontface.names}")
            except Exception as e:
                errors.append(f"frontface: {type(e).__name__}: {e}")
        else:
            errors.append(f"frontface: weights file not found at {YOLO_FRONTFACE}")

        yolo_available = model_endface is not None or model_frontface is not None
        if errors:
            yolo_error = "; ".join(errors)
            print(f"[GST] YOLO partial/failed load: {yolo_error}")
        if not yolo_available:
            print("[GST] YOLO unavailable (no models loaded). Streaming raw frames.")

    cap = cv2.VideoCapture(GST_STR, cv2.CAP_GSTREAMER)
    if not cap.isOpened():
        print("[GST] GStreamer pipeline failed - falling back to direct RTSP.")
        cap = cv2.VideoCapture(RTSP_URL)

    detect_frame_count = 0
    last_detect_error = ""

    def _annotate(frm, model, cls_names, line_side):
        """line_side: 'right' or 'left' - which edge of the box to draw the
        vertical marker line on. This used to be picked via `clsid == 0`, which
        worked when a single 2-class model produced clsid 0 (frontface) or 1
        (endface). Now each model is loaded separately with exactly one class,
        so clsid is *always* 0 for both models - the old check could no longer
        distinguish frontface from endface and always drew on the same side.
        Passing the side explicitly per model call fixes that."""
        nonlocal detect_frame_count, last_detect_error
        if model is None:
            return frm
        try:
            preds = list(model.predict(source=[frm], conf=0.6, save=False, verbose=False))
            n = len(preds[0])
            detect_frame_count += 1
            if detect_frame_count % 60 == 0:
                print(f"[GST] predict() ran, {n} detections this frame, "
                      f"model classes={model.names}")
            if n == 0:
                return frm
            for i in range(n):
                box   = preds[0].boxes[i]
                clsid = int(box.cls.cpu().numpy()[0])
                conf  = float(box.conf.cpu().numpy()[0])
                bb    = box.xyxy.cpu().numpy()[0]
                h, w  = frm.shape[:2]
                cv2.rectangle(frm, (int(bb[0]), int(bb[1]) - 20), (int(bb[2]), int(bb[1])), (0, 0, 0), -1)
                cv2.rectangle(frm, (int(bb[0]), int(bb[1])), (int(bb[2]), int(bb[3])), (30, 120, 255), 2)
                if line_side == "right":
                    cv2.line(frm, (int(bb[2]), 0), (int(bb[2]), h), (0, 220, 100), 3)
                else:
                    cv2.line(frm, (int(bb[0]), 0), (int(bb[0]), h), (0, 220, 100), 3)
                cls_label = cls_names[clsid] if clsid < len(cls_names) else f"cls{clsid}"
                cv2.putText(frm, f"{cls_label} {round(conf * 100, 1)}%",
                            (int(bb[0]), int(bb[1]) - 5), cv2.FONT_HERSHEY_PLAIN, 1, (255, 255, 255), 1)
        except Exception as ex:
            import traceback
            last_detect_error = f"{type(ex).__name__}: {ex}"
            print(f"[GST] Detection error: {last_detect_error}")
            traceback.print_exc()
        return frm

    def _draw_yolo_status(frm):
        """Always show YOLO status in a corner so failures are visible in the browser,
        not just buried in a terminal you might not be watching."""
        if model_endface is not None and model_frontface is not None:
            msg, color = "YOLO: OK (endface + frontface)", (0, 220, 100)
        elif yolo_available:
            loaded = "endface" if model_endface is not None else "frontface"
            msg, color = f"YOLO PARTIAL: only {loaded} loaded - {yolo_error}", (0, 165, 255)
        else:
            msg, color = f"YOLO OFF: {yolo_error or 'unknown error'}", (0, 0, 255)
        cv2.putText(frm, msg, (10, frm.shape[0] - 12),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1, cv2.LINE_AA)
        return frm

    while True:
        ret, frame = cap.read()
        if not ret:
            print("[GST] Stream lost. Reconnecting...")
            try:
                cap.release()
            except Exception:
                pass
            time.sleep(2)
            cap = cv2.VideoCapture(GST_STR, cv2.CAP_GSTREAMER)
            if not cap.isOpened():
                cap = cv2.VideoCapture(RTSP_URL)
            continue
        frame = cv2.resize(frame, (1140, 640))
        if yolo_available:
            frame = _annotate(frame, model_frontface, cls_frontface, "right")
            frame = _annotate(frame, model_endface,   cls_endface,   "left")
        frame = _draw_yolo_status(frame)
        frame = draw_video_overlay(frame, _state_payload())
        ok, jpeg = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
        if ok:
            with _latest_gst_frame_lock:
                _latest_gst_frame = jpeg.tobytes()

# ------------------------------------------------------------------
# THREAD 3 — BILLET DATA REFRESH
# ------------------------------------------------------------------
def billet_refresh_loop() -> None:
    while True:
        try:
            conn   = sqlite3.connect(SQLITE_DB, check_same_thread=False, timeout=5)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM shared_variable")
            result = cursor.fetchall()
            c_time = datetime.now().time()
            s_date = get_previous_date() if dtime(0, 0) <= c_time <= dtime(5, 59) else datetime.now().strftime("%Y-%m-%d")
            s_time = s_date + "T06:00"
            e_time = datetime.now().strftime("%Y-%m-%dT%H:%M")
            cursor.execute("SELECT COUNT(b_length) FROM full_data WHERE date BETWEEN ? AND ?", (s_time, e_time))
            bc = cursor.fetchone()
            conn.close()
            with _billet_lock:
                if result:
                    _billet_state["billet_len"] = result[1][1] if len(result) > 1 else 0
                    _billet_state["len_6"]      = result[2][1] if len(result) > 2 else 0
                    _billet_state["len_9"]      = result[3][1] if len(result) > 3 else 0
                    _billet_state["len_12"]     = result[4][1] if len(result) > 4 else 0
                    _billet_state["uncat"]      = result[5][1] if len(result) > 5 else 0
                    _billet_state["final_len"]  = result[6][1] if len(result) > 6 else 0
                _billet_state["billet_count"] = bc[0] if bc else 0
        except Exception as e:
            print(f"[BILLET] DB read error: {e}")
        time.sleep(5)

# ------------------------------------------------------------------
# FLASK APPLICATION
# ------------------------------------------------------------------
app = Flask(__name__)

@app.route("/video_feed")
def video_feed():
    def generate():
        while True:
            with _latest_frame_lock:
                frame = _latest_frame
            if frame is None:
                time.sleep(0.03)
                continue
            yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + frame + b"\r\n"
            time.sleep(0.033)
    return Response(stream_with_context(generate()), mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/stream_feed")
def stream_feed():
    def generate():
        while True:
            with _latest_gst_frame_lock:
                frame = _latest_gst_frame
            if frame is None:
                time.sleep(0.04)
                continue
            yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + frame + b"\r\n"
            time.sleep(0.04)
    return Response(stream_with_context(generate()), mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/events")
def events():
    q: queue.Queue = queue.Queue(maxsize=100)
    with _sse_lock:
        _sse_clients.append(q)

    def generate():
        try:
            yield "event: metrics\ndata: " + json.dumps(_state_payload()) + "\n\n"
            while True:
                try:
                    yield q.get(timeout=20)
                except queue.Empty:
                    yield ": keepalive\n\n"
        except GeneratorExit:
            pass
        finally:
            with _sse_lock:
                try:
                    _sse_clients.remove(q)
                except ValueError:
                    pass

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Access-Control-Allow-Origin": "*"},
    )

@app.route("/api/state")
def api_state():
    return jsonify({"status": "ok", "items": ITEMS, "state": _state_payload()})

@app.route("/api/billet")
def api_billet():
    with _billet_lock:
        b = dict(_billet_state)
    return jsonify({"status": "ok", "billet": b})

@app.route("/api/blade_logs")
def api_blade_logs():
    rows = read_csv_logs(BLADE_CSV_PATH, BLADE_HEADERS)
    return jsonify({"status": "ok", "logs": rows})

@app.route("/api/item_logs")
def api_item_logs():
    rows = read_csv_logs(ITEM_CSV_PATH, ITEM_HEADERS)
    return jsonify({"status": "ok", "logs": rows})

@app.route("/api/logs")
def api_logs():
    rows = read_csv_logs(BLADE_CSV_PATH, BLADE_HEADERS)
    return jsonify({"status": "ok", "logs": rows})

# ------------------------------------------------------------------
# NEW ROUTE: Change item without resetting blade count
# ------------------------------------------------------------------
@app.route("/change_item", methods=["POST"])
def change_item():
    """Switch current_item immediately without touching blade_count or any counters."""
    global _change_item_flag, _change_item_new_item
    data     = request.get_json(silent=True) or {}
    new_item = str(data.get("new_item", "")).strip()
    if new_item not in ITEMS:
        return jsonify({"status": "error", "message": f"Unknown item: {new_item}"}), 400

    # Flush the outgoing item's current count into the active Item Wise
    # session row before switching, so the sheet reflects it immediately.
    with _state_lock:
        current_item  = str(_state.get("current_item", ITEMS[0]))
        current_count = int(_state.get("item_counts", {}).get(current_item, 0))
    try:
        update_item_session(current_item, current_count)
    except Exception as e:
        print(f"[WARN] update_item_session on change_item failed: {e}")

    with _change_item_flag_lock:
        _change_item_flag     = True
        _change_item_new_item = new_item
    return jsonify({"status": "ok", "new_item": new_item})

@app.route("/reset_item", methods=["POST"])
def reset_item():
    data      = request.get_json(silent=True) or {}
    next_item = str(data.get("next_item", "")).strip()
    with _state_lock:
        current_item   = str(_state.get("current_item", ITEMS[0]))
        current_count  = int(_state.get("item_counts", {}).get(current_item, 0))
        blade_count    = int(_state.get("blade_count", 0))
    try:
        update_item_session(current_item, current_count)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    with _state_lock:
        _state["item_counts"][current_item]  = 0
        _state["item_session_start_time"]    = _now_str()
        if next_item in ITEMS:
            _state["current_item"] = next_item
            _state["item_counts"].setdefault(next_item, 0)
            _state["blade_item_counts"].setdefault(next_item, 0)
    save_runtime_state()
    write_live_item_state()
    _broadcast_event("item_reset", json.dumps(_state_payload()))
    return jsonify({"status": "ok", "logged": True, "blade_count": blade_count, "state": _state_payload()})

@app.route("/reset_blade", methods=["POST"])
def reset_blade():
    global _reset_blade_flag, _reset_blade_next_item
    data      = request.get_json(silent=True) or {}
    next_item = str(data.get("next_item", "")).strip()
    with _state_lock:
        blade_count   = int(_state.get("blade_count", 0))
        blade_start   = str(_state.get("blade_start_time", _now_str()))
        current_item  = str(_state.get("current_item", ITEMS[0]))
        current_count = int(_state.get("item_counts", {}).get(current_item, 0))
        breakdown     = dict(_state.get("blade_item_counts", {}))
    try:
        entry = append_blade_report(blade_count, blade_start, current_item, breakdown)
        # Finalize the current item's count into its session row before
        # starting a fresh blank row for the new blade session.
        update_item_session(current_item, current_count)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    if next_item not in ITEMS:
        next_item = current_item
    try:
        create_new_item_session_row()
    except Exception as e:
        print(f"[WARN] create_new_item_session_row failed: {e}")
    with _reset_blade_flag_lock:
        _reset_blade_flag      = True
        _reset_blade_next_item = next_item
    return jsonify({"status": "ok", "logged": True, "entry": entry, "next_item": next_item})

@app.route("/reset", methods=["POST"])
def reset_old():
    return reset_blade()

@app.route("/download/blade_csv")
def download_blade_csv():
    ensure_csv_header(BLADE_CSV_PATH, BLADE_HEADERS)
    return send_file(BLADE_CSV_PATH, as_attachment=True, download_name="blade_reset_report.csv")

@app.route("/download/item_csv")
def download_item_csv():
    ensure_csv_header(ITEM_CSV_PATH, ITEM_HEADERS)
    return send_file(ITEM_CSV_PATH, as_attachment=True, download_name="item_wise_reset_report.csv")

@app.route("/download/combined_excel")
def download_combined_excel():
    path = build_combined_excel()
    return send_file(path, as_attachment=True, download_name="msp_hotsow_combined_report.xlsx")

@app.route("/download/blade_live_csv")
def download_blade_live_csv():
    write_live_blade_state()
    return send_file(BLADE_LIVE_CSV_PATH, as_attachment=True, download_name="blade_live_session.csv")

@app.route("/download/item_live_csv")
def download_item_live_csv():
    write_live_item_state()
    return send_file(ITEM_LIVE_CSV_PATH, as_attachment=True, download_name="item_live_session.csv")

@app.route("/download/csv")
def download_csv():
    return download_blade_csv()

@app.route("/download/excel")
def download_excel():
    return download_combined_excel()

# ------------------------------------------------------------------
# HTML TEMPLATES
# ------------------------------------------------------------------
VISION_HTML = r"""
<!DOCTYPE html><html><head><meta charset="UTF-8"><title>MSP Vision Feed</title>
<style>body{margin:0;background:#050810;color:white;font-family:Arial;height:100vh;display:flex;flex-direction:column}
.hdr{height:54px;background:#0f172a;display:flex;align-items:center;justify-content:space-between;padding:0 22px}
.hdr a{color:white;text-decoration:none;background:#1e40af;padding:8px 14px;border-radius:6px}
.body{flex:1;display:flex;align-items:center;justify-content:center}
.body img{width:100%;height:100%;object-fit:contain}</style></head>
<body><div class="hdr"><b>MSP YOLO Billet Detection – Live Feed</b><a href="/">Main Dashboard</a></div>
<div class="body"><img src="/stream_feed"></div></body></html>
"""

DASHBOARD_HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>MSP Hotsow 1 – Cut Counter</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@700;800&display=swap" rel="stylesheet">
<style>
/* ── Design tokens ── */
:root {
  --primary:      #1E40AF;
  --primary-mid:  #2563EB;
  --primary-lt:   #EFF6FF;
  --orange:       #EA580C;
  --orange-lt:    #FFF7ED;
  --orange-ring:  #FDBA74;
  --green:        #059669;
  --green-lt:     #ECFDF5;
  --red:          #DC2626;
  --red-lt:       #FEF2F2;
  --amber:        #D97706;
  --purple:       #7C3AED;
  --purple-lt:    #F5F3FF;
  --bg:           #F1F5F9;
  --card:         #FFFFFF;
  --border:       #E2E8F0;
  --border-mid:   #CBD5E1;
  --text:         #0F172A;
  --text-mid:     #334155;
  --muted:        #64748B;
  --muted-lt:     #94A3B8;
  --shadow-sm:    0 1px 3px rgba(0,0,0,.07), 0 1px 8px rgba(0,0,0,.05);
  --shadow:       0 2px 8px rgba(0,0,0,.08), 0 4px 20px rgba(0,0,0,.06);
  --shadow-lg:    0 8px 32px rgba(0,0,0,.12);
  --radius:       14px;
  --radius-sm:    8px;
}

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: 'Inter', Arial, sans-serif;
  background: var(--bg);
  color: var(--text);
  height: 100vh;
  overflow: hidden;
  font-size: 13px;
}

/* ── Header ── */
.header {
  height: 56px;
  background: linear-gradient(135deg, #1E3A8A 0%, #2563EB 60%, #3B82F6 100%);
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: 0 20px;
  box-shadow: 0 2px 12px rgba(30,64,175,.35);
  position: relative;
  z-index: 10;
}
.header-left { display: flex; align-items: center; gap: 12px; }
.header-logo {
  width: 34px; height: 34px;
  background: rgba(255,255,255,.18);
  border-radius: 8px;
  display: flex; align-items: center; justify-content: center;
  font-size: 18px;
}
.header-title { font-size: 17px; font-weight: 800; color: #fff; letter-spacing: .5px; line-height: 1.1; }
.header-sub { font-size: 10px; color: rgba(255,255,255,.65); font-weight: 500; letter-spacing: .8px; }
.header-right { display: flex; align-items: center; gap: 10px; }
.clock-badge {
  font-family: 'JetBrains Mono', monospace;
  font-size: 14px; font-weight: 700;
  color: rgba(255,255,255,.9);
  background: rgba(255,255,255,.12);
  border: 1px solid rgba(255,255,255,.2);
  border-radius: 7px;
  padding: 5px 12px;
}
.hlink {
  background: rgba(255,255,255,.15);
  border: 1px solid rgba(255,255,255,.25);
  color: #fff;
  border-radius: 7px;
  padding: 6px 13px;
  font-size: 12px;
  font-weight: 600;
  text-decoration: none;
  transition: background .15s;
}
.hlink:hover { background: rgba(255,255,255,.28); }

/* ── Main 3-col grid ── */
.main {
  height: calc(100vh - 56px);
  display: grid;
  grid-template-columns: 1fr 370px 320px;
  gap: 10px;
  padding: 10px;
}

/* ── Generic card ── */
.card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
}

/* ── Video column ── */
.video-col { display: flex; flex-direction: column; }
.video-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  overflow: hidden;
  flex: 1;
  display: flex;
  flex-direction: column;
}
.video-header {
  background: #0F172A;
  color: #fff;
  padding: 10px 14px;
  display: flex;
  justify-content: space-between;
  align-items: center;
}
.video-header-left { display: flex; align-items: center; gap: 8px; }
.live-dot {
  width: 8px; height: 8px;
  background: #22C55E;
  border-radius: 50%;
  animation: pulse-dot 1.4s ease infinite;
}
@keyframes pulse-dot {
  0%,100% { opacity: 1; transform: scale(1); }
  50%      { opacity: .5; transform: scale(.75); }
}
.video-header-title { font-size: 11px; font-weight: 700; letter-spacing: 1px; color: #E2E8F0; }
.frame-badge {
  font-family: 'JetBrains Mono', monospace;
  font-size: 10px;
  color: #64748B;
  background: #1E293B;
  padding: 3px 8px;
  border-radius: 5px;
}
.video-body {
  flex: 1;
  background: #020617;
  display: flex;
  align-items: center;
  justify-content: center;
  overflow: hidden;
}
.video-body img { width: 100%; height: 100%; object-fit: contain; }
.video-foot {
  background: #F8FAFC;
  border-top: 1px solid var(--border);
  padding: 6px 12px;
  font-size: 10px;
  color: var(--muted);
  font-family: 'JetBrains Mono', monospace;
}

/* ── Middle column ── */
.mid-col { display: flex; flex-direction: column; gap: 8px; overflow: hidden; }

/* Blade count card */
.blade-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  border-left: 4px solid var(--orange);
  padding: 14px 16px;
}
.card-eyebrow {
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 1.2px;
  text-transform: uppercase;
  color: var(--muted);
  margin-bottom: 6px;
}
.blade-num-row { display: flex; align-items: flex-end; gap: 10px; }
.blade-num {
  font-family: 'JetBrains Mono', monospace;
  font-size: 68px;
  font-weight: 800;
  line-height: .9;
  color: var(--orange);
  transition: color .2s;
}
.blade-num.flash { animation: blade-flash .4s ease; }
@keyframes blade-flash {
  0%   { color: var(--orange); transform: scale(1); }
  40%  { color: #DC2626;       transform: scale(1.06); }
  100% { color: var(--orange); transform: scale(1); }
}
.blade-label-unit {
  font-size: 12px; font-weight: 700;
  color: var(--muted-lt);
  margin-bottom: 10px;
  text-transform: uppercase;
  letter-spacing: .5px;
}
.blade-meta { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 10px; }
.meta-box {
  background: #F8FAFC;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  padding: 9px 11px;
}
.meta-label { font-size: 9px; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; color: var(--muted); }
.meta-val {
  font-family: 'JetBrains Mono', monospace;
  font-size: 13px;
  font-weight: 700;
  color: var(--text-mid);
  margin-top: 3px;
}

/* Item card */
.item-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  border-left: 4px solid var(--green);
  padding: 14px 16px;
}
.styled-select {
  width: 100%;
  padding: 10px 12px;
  border: 1.5px solid var(--border-mid);
  border-radius: var(--radius-sm);
  font-size: 14px;
  font-weight: 700;
  font-family: 'Inter', sans-serif;
  background: #F8FAFC;
  color: var(--text);
  cursor: pointer;
  margin-top: 8px;
  outline: none;
  transition: border-color .15s;
}
.styled-select:focus { border-color: var(--primary-mid); background: #fff; }

/* 2-button row for item actions */
.btn-row-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 10px; }
.btn-row   { display: grid; grid-template-columns: 1fr; gap: 8px; margin-top: 10px; }

.btn {
  border: none;
  border-radius: var(--radius-sm);
  padding: 10px 13px;
  font-size: 12px;
  font-weight: 700;
  font-family: 'Inter', sans-serif;
  cursor: pointer;
  letter-spacing: .4px;
  transition: all .15s;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  gap: 5px;
}
.btn-blue   { background: #DBEAFE; color: var(--primary);  border: 1px solid #BFDBFE; }
.btn-blue:hover   { background: var(--primary-mid); color: #fff; }
.btn-green  { background: #D1FAE5; color: var(--green);   border: 1px solid #A7F3D0; }
.btn-green:hover  { background: var(--green); color: #fff; }
.btn-purple { background: var(--purple-lt); color: var(--purple); border: 1px solid #DDD6FE; }
.btn-purple:hover { background: var(--purple); color: #fff; }
.btn-orange { background: #FFF7ED; color: var(--orange);  border: 1px solid #FED7AA; }
.btn-orange:hover { background: var(--orange); color: #fff; }
.btn-red    { background: #FEE2E2; color: var(--red);     border: 1px solid #FECACA; }
.btn-red:hover    { background: var(--red); color: #fff; }
.btn-ghost  { background: #F1F5F9; color: var(--muted);   border: 1px solid var(--border); }
.btn-ghost:hover  { background: #E2E8F0; }
.btn-danger { background: var(--red); color: #fff; border: 1px solid var(--red); }
.btn-danger:hover { background: #B91C1C; }

.blade-reset-btn {
  width: 100%;
  padding: 13px;
  font-size: 13px;
  border-radius: var(--radius);
  background: linear-gradient(135deg, #991B1B 0%, #DC2626 100%);
  color: #fff;
  border: none;
  cursor: pointer;
  font-weight: 800;
  font-family: 'Inter', sans-serif;
  letter-spacing: .6px;
  box-shadow: 0 2px 8px rgba(220,38,38,.3);
  transition: all .18s;
  display: flex;
  align-items: center;
  justify-content: center;
  gap: 8px;
}
.blade-reset-btn:hover { background: linear-gradient(135deg, #7F1D1D 0%, #B91C1C 100%); box-shadow: 0 4px 16px rgba(220,38,38,.4); }

/* change-item hint pill */
.change-hint {
  display: inline-block;
  font-size: 9px;
  font-weight: 700;
  letter-spacing: .8px;
  text-transform: uppercase;
  background: #F5F3FF;
  color: var(--purple);
  border: 1px solid #DDD6FE;
  border-radius: 999px;
  padding: 2px 8px;
  margin-left: 6px;
  vertical-align: middle;
}

.item-stats { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 10px; }
.stat-box {
  background: var(--green-lt);
  border: 1px solid #A7F3D0;
  border-radius: var(--radius-sm);
  padding: 9px 11px;
}
.stat-box .meta-val { color: var(--green); font-size: 22px; }

/* Detection metrics */
.metrics-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  padding: 12px 14px;
}
.metrics-grid { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 8px; margin-top: 8px; }
.metric-box {
  background: #F8FAFC;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  padding: 8px 10px;
  text-align: center;
}
.metric-val {
  font-family: 'JetBrains Mono', monospace;
  font-size: 18px;
  font-weight: 800;
  color: var(--primary-mid);
  margin-top: 3px;
}

/* ── Right column ── */
.right-col { display: flex; flex-direction: column; gap: 8px; overflow: hidden; }
.report-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  display: flex;
  flex-direction: column;
  flex: 1;
  min-height: 0;
  overflow: hidden;
}
.tabs-bar {
  display: flex;
  gap: 6px;
  padding: 8px 8px 0;
  background: #F8FAFC;
  border-bottom: 1px solid var(--border);
}
.tab {
  flex: 1;
  padding: 8px 4px;
  border: 1px solid var(--border);
  border-bottom: none;
  border-radius: 6px 6px 0 0;
  background: #fff;
  cursor: pointer;
  font-weight: 700;
  font-family: 'Inter', sans-serif;
  font-size: 11px;
  color: var(--muted);
  transition: all .15s;
}
.tab.active { background: var(--primary-mid); color: #fff; border-color: var(--primary-mid); }
.search-bar {
  display: flex;
  gap: 6px;
  padding: 8px;
  background: #F8FAFC;
  border-bottom: 1px solid var(--border);
}
.search-input {
  flex: 1;
  padding: 7px 10px;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  font-size: 12px;
  font-family: 'Inter', sans-serif;
  outline: none;
  background: #fff;
}
.search-input:focus { border-color: var(--primary-mid); }
.date-filter-bar {
  display: flex;
  align-items: center;
  gap: 6px;
  padding: 0 8px 8px;
  background: #F8FAFC;
  border-bottom: 1px solid var(--border);
}
.date-input {
  flex: 1;
  min-width: 0;
  padding: 6px 8px;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  font-size: 11px;
  font-family: 'Inter', sans-serif;
  outline: none;
  background: #fff;
  color: var(--text);
}
.date-input:focus { border-color: var(--primary-mid); }
.date-sep { font-size: 10px; color: var(--muted); font-weight: 600; }
.table-wrap { overflow: auto; flex: 1; }
table { width: 100%; border-collapse: collapse; font-size: 11px; }
th {
  position: sticky; top: 0;
  background: #F1F5F9;
  color: var(--muted);
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 1px;
  text-transform: uppercase;
  padding: 8px 8px;
  border-bottom: 1px solid var(--border);
  white-space: nowrap;
}
td { padding: 8px 8px; border-bottom: 1px solid #F1F5F9; vertical-align: top; }
tr:hover td { background: #F8FAFC; }
.id-badge {
  display: inline-block;
  font-family: 'JetBrains Mono', monospace;
  font-size: 10px;
  font-weight: 700;
  background: var(--primary-lt);
  color: var(--primary);
  border-radius: 4px;
  padding: 1px 5px;
}
.count-bold {
  font-family: 'JetBrains Mono', monospace;
  font-weight: 800;
  font-size: 13px;
  color: var(--orange);
}
.item-cell {
  font-family: 'JetBrains Mono', monospace;
  font-weight: 700;
  font-size: 12px;
  color: var(--text-mid);
}
.item-cell.nonzero { color: var(--orange); }
.exports-row {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 8px;
  padding: 8px;
  border-top: 1px solid var(--border);
  background: #F8FAFC;
}

/* Breakdown card */
.breakdown-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  padding: 12px 14px;
  max-height: 200px;
  overflow: hidden;
  display: flex;
  flex-direction: column;
}
.breakdown-list { overflow: auto; flex: 1; margin-top: 8px; }
.brow {
  display: flex;
  justify-content: space-between;
  align-items: center;
  padding: 5px 0;
  border-bottom: 1px solid #F1F5F9;
  font-size: 12px;
}
.brow:last-child { border-bottom: none; }
.brow-name { color: var(--text-mid); font-weight: 500; }
.brow-badge {
  font-family: 'JetBrains Mono', monospace;
  font-size: 11px;
  font-weight: 700;
  background: var(--orange-lt);
  color: var(--orange);
  border: 1px solid var(--orange-ring);
  border-radius: 999px;
  padding: 1px 8px;
  min-width: 28px;
  text-align: center;
}
.brow-badge.nonzero { background: var(--orange); color: #fff; border-color: var(--orange); }

/* ── Modal ── */
.modal-overlay {
  position: fixed; inset: 0;
  background: rgba(15,23,42,.55);
  backdrop-filter: blur(4px);
  display: flex;
  align-items: center;
  justify-content: center;
  z-index: 9000;
  opacity: 0;
  pointer-events: none;
  transition: opacity .2s;
}
.modal-overlay.open { opacity: 1; pointer-events: auto; }
.modal-box {
  background: #fff;
  border-radius: 18px;
  box-shadow: var(--shadow-lg);
  padding: 28px 28px 24px;
  width: 380px;
  max-width: calc(100vw - 32px);
  transform: translateY(16px) scale(.97);
  transition: transform .22s cubic-bezier(.34,1.56,.64,1);
}
.modal-overlay.open .modal-box { transform: translateY(0) scale(1); }
.modal-icon { font-size: 30px; margin-bottom: 12px; }
.modal-title { font-size: 18px; font-weight: 800; color: var(--text); margin-bottom: 8px; }
.modal-desc { font-size: 13px; color: var(--muted); line-height: 1.55; margin-bottom: 18px; }
.modal-desc strong { color: var(--text); }
.modal-label {
  display: block;
  font-size: 11px; font-weight: 700; letter-spacing: .8px;
  text-transform: uppercase; color: var(--muted);
  margin-bottom: 6px;
}
.modal-select {
  width: 100%;
  padding: 11px 13px;
  border: 1.5px solid var(--border-mid);
  border-radius: var(--radius-sm);
  font-size: 14px; font-weight: 700;
  font-family: 'Inter', sans-serif;
  background: #F8FAFC;
  color: var(--text);
  outline: none;
  margin-bottom: 20px;
}
.modal-select:focus { border-color: var(--primary-mid); }
.modal-actions { display: flex; gap: 10px; }
.modal-actions .btn { flex: 1; padding: 12px; font-size: 13px; }

/* ── Toast ── */
.toast {
  position: fixed;
  right: 18px; bottom: 18px;
  background: #0F172A;
  color: #fff;
  border-radius: 10px;
  padding: 11px 18px;
  font-size: 13px; font-weight: 600;
  box-shadow: 0 4px 20px rgba(0,0,0,.25);
  opacity: 0; transform: translateY(8px);
  transition: opacity .2s, transform .2s;
  z-index: 9999;
  max-width: 320px;
}
.toast.show { opacity: 1; transform: translateY(0); }
.toast.success { background: var(--green); }
.toast.error   { background: var(--red); }
.toast.info    { background: var(--purple); }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #CBD5E1; border-radius: 99px; }
</style>
</head>
<body>

<!-- ── Header ── -->
<header class="header">
  <div class="header-left">
    <div class="header-logo">⚙️</div>
    <div>
      <div class="header-title">HOTSOW 1 – Cut Counter</div>
      <div class="header-sub">MSP RAIGARH · BLADE + ITEM TRACKING</div>
    </div>
  </div>
  <div class="header-right">
    <span class="clock-badge" id="clock">--:--:--</span>
    <a class="hlink" href="/vision" target="_blank">Vision Feed ↗</a>
  </div>
</header>

<!-- ── Main grid ── -->
<main class="main">

  <!-- Video column -->
  <div class="video-col">
    <div class="video-card">
      <div class="video-header">
        <div class="video-header-left">
          <div class="live-dot"></div>
          <span class="video-header-title">CAMERA FEED – CUT DETECTION</span>
        </div>
        <span class="frame-badge" id="frameCt">FRAME 000000</span>
      </div>
      <div class="video-body">
        <img src="/stream_feed" onerror="this.outerHTML='<div style=color:#475569;padding:20px>Waiting for camera stream…</div>'">
      </div>
      <div class="video-foot">Source: {{ rtsp_url }}</div>
    </div>
  </div>

  <!-- Middle column -->
  <div class="mid-col">

    <!-- Blade count card -->
    <div class="blade-card">
      <div class="card-eyebrow">🔪 Running Blade Count</div>
      <div class="blade-num-row">
        <div class="blade-num" id="bladeCount">0</div>
        <div class="blade-label-unit">cuts</div>
      </div>
      <div class="blade-meta">
        <div class="meta-box">
          <div class="meta-label">Blade Started</div>
          <div class="meta-val" id="bladeStart" style="font-size:11px">–</div>
        </div>
        <div class="meta-box">
          <div class="meta-label">Last Cut</div>
          <div class="meta-val" id="lastCut">–</div>
        </div>
      </div>
    </div>

    <!-- Current item card -->
    <div class="item-card">
      <div class="card-eyebrow">
        📦 Current Running Item
        <span class="change-hint">no-reset switch available</span>
      </div>
      <select id="itemSelect" class="styled-select"></select>

      <!-- Two buttons side by side -->
      <div class="btn-row-2">
        <button id="changeItemBtn" class="btn btn-purple" title="Switch item immediately — blade count stays unchanged">
          ⇄ Change Item
        </button>
        <button id="resetItemBtn" class="btn btn-green" title="Save item report and reset item count">
          ↺ Reset Item
        </button>
      </div>

      <div class="item-stats">
        <div class="stat-box">
          <div class="meta-label">Item Count</div>
          <div class="meta-val" id="itemCount">0</div>
        </div>
        <div class="meta-box">
          <div class="meta-label">Item Started</div>
          <div class="meta-val" id="itemStart" style="font-size:11px">–</div>
        </div>
      </div>
    </div>

    <!-- Reset blade button -->
    <button class="blade-reset-btn" id="resetBladeBtn">
      ⚡ Reset Blade &amp; Save Report
    </button>

    <!-- Detection metrics -->
    <div class="metrics-card">
      <div class="card-eyebrow">📡 Detection Metrics</div>
      <div class="metrics-grid">
        <div class="metric-box">
          <div class="meta-label">Bright</div>
          <div class="metric-val" id="mBright">0</div>
        </div>
        <div class="metric-box">
          <div class="meta-label">Motion</div>
          <div class="metric-val" id="mMotion">0</div>
        </div>
        <div class="metric-box">
          <div class="meta-label">Score</div>
          <div class="metric-val" id="mScore">0/7</div>
        </div>
      </div>
    </div>

  </div><!-- /mid-col -->

  <!-- Right column -->
  <div class="right-col">

    <!-- Report table -->
    <div class="report-card">
      <div class="tabs-bar">
        <button class="tab active" id="bladeTab">Blade Resets</button>
        <button class="tab" id="itemTab">Item Wise</button>
      </div>
      <div class="search-bar">
        <input class="search-input" id="search" placeholder="Search…">
        <button class="btn btn-blue" id="refreshBtn" style="padding:7px 12px">↺</button>
      </div>
      <div class="date-filter-bar">
        <input type="date" class="date-input" id="dateFrom" title="From date">
        <span class="date-sep">to</span>
        <input type="date" class="date-input" id="dateTo" title="To date">
        <button class="btn btn-blue" id="clearDateBtn" style="padding:6px 10px" title="Clear date filter">✕</button>
      </div>
      <div class="table-wrap">
        <table><thead id="thead"></thead><tbody id="tbody"></tbody></table>
      </div>
      <div class="exports-row">
        <a class="btn btn-blue" id="csvLink" href="/download/blade_csv">↓ CSV</a>
        <a class="btn btn-green" id="excelLink" href="/download/combined_excel">↓ Excel (both sheets)</a>
      </div>
    </div>

    <!-- Blade item breakdown -->
    <div class="breakdown-card">
      <div class="card-eyebrow">📊 Blade Breakdown (Current)</div>
      <div class="breakdown-list" id="itemBreakdown"></div>
    </div>

  </div><!-- /right-col -->
</main>

<!-- ── Blade Reset Modal ── -->
<div id="bladeResetModal" class="modal-overlay">
  <div class="modal-box">
    <div class="modal-icon">⚡</div>
    <div class="modal-title">Reset Blade Counter</div>
    <p class="modal-desc">
      The current blade count will be <strong>saved to the report</strong> and reset to <strong>0</strong>.
      Select the item that will run on the new blade.
    </p>
    <label class="modal-label" for="bladeResetItemSelect">Next Running Item</label>
    <select id="bladeResetItemSelect" class="modal-select"></select>
    <div class="modal-actions">
      <button id="cancelBladeResetBtn" class="btn btn-ghost">Cancel</button>
      <button id="confirmBladeResetBtn" class="btn btn-danger">Save &amp; Reset</button>
    </div>
  </div>
</div>

<!-- ── Change Item Modal (no reset) ── -->
<div id="changeItemModal" class="modal-overlay">
  <div class="modal-box">
    <div class="modal-icon">⇄</div>
    <div class="modal-title">Change Current Item</div>
    <p class="modal-desc">
      Switch to a different item <strong>immediately</strong>.<br>
      The blade count and all existing counts are <strong>not affected</strong>.<br>
      Future cuts will be credited to the new item.
    </p>
    <label class="modal-label" for="changeItemSelect">Switch To</label>
    <select id="changeItemSelect" class="modal-select"></select>
    <div class="modal-actions">
      <button id="cancelChangeItemBtn" class="btn btn-ghost">Cancel</button>
      <button id="confirmChangeItemBtn" class="btn btn-purple">Switch Item</button>
    </div>
  </div>
</div>

<!-- ── Toast ── -->
<div id="toast" class="toast">Done</div>

<script>
let items = [], state = {}, bladeLogs = [], itemLogs = [], activeTab = 'blade';
const $ = id => document.getElementById(id);

/* Toast */
function toast(msg, type = '') {
  const t = $('toast');
  t.textContent = msg;
  t.className = 'toast show' + (type ? ' ' + type : '');
  setTimeout(() => t.className = 'toast', 3200);
}

/* Clock */
function clock() { $('clock').textContent = new Date().toTimeString().slice(0, 8); }
setInterval(clock, 1000); clock();

/* Populate all selects once */
function populateSelects(items, currentItem) {
  ['itemSelect', 'bladeResetItemSelect', 'changeItemSelect'].forEach(id => {
    const sel = $(id);
    if (sel.options.length === 0) {
      items.forEach(it => {
        const o = document.createElement('option');
        o.value = it; o.textContent = it;
        sel.appendChild(o);
      });
    }
    sel.value = currentItem || items[0];
  });
}

/* Load initial state */
async function loadState() {
  const r = await fetch('/api/state');
  const d = await r.json();
  items = d.items || [];
  state = d.state || {};
  populateSelects(items, state.current_item);
  renderState();
}

/* Render live state */
function renderState() {
  $('bladeCount').textContent = state.blade_count || 0;
  $('bladeStart').textContent = state.blade_start_time || '–';
  $('itemCount').textContent  = state.current_item_count || 0;
  $('itemStart').textContent  = state.item_session_start_time || '–';
  $('frameCt').textContent    = 'FRAME ' + String(state.frame_id || 0).padStart(6, '0');
  $('mBright').textContent    = Number(state.bright || 0).toLocaleString();
  $('mMotion').textContent    = Number(state.motion || 0).toLocaleString();
  $('mScore').textContent     = (state.score || 0) + '/7';

  /* Keep item selects in sync */
  if (state.current_item) {
    $('itemSelect').value = state.current_item;
  }

  /* Blade breakdown */
  const br = state.blade_item_counts || {};
  $('itemBreakdown').innerHTML = items.map(it => {
    const cnt = br[it] || 0;
    return `<div class="brow">
      <span class="brow-name">${it}</span>
      <span class="brow-badge ${cnt > 0 ? 'nonzero' : ''}">${cnt}</span>
    </div>`;
  }).join('');
}

/* Load log tables */
async function loadLogs() {
    const [b, i] = await Promise.all([
        fetch('/api/blade_logs').then(r => r.json()),
        fetch('/api/item_logs').then(r => r.json()),
    ]);

    bladeLogs = b.logs || [];
    itemLogs = i.logs || [];

    renderLogs();
}

/* Item Wise table columns */
const ITEM_COLUMNS = [
    "MB 150",
    "MB 200",
    "MB 250",
    "MB 300",
    "MB 350",
    "MB 400",
    "MC 150",
    "MC 175",
    "MC 200",
    "MC 250",
    "MC 300",
    "Angle 130*10/12",
    "Angle 150*10/12"
];

/* Render active tab */
function renderLogs() {

    const q = ($('search').value || '').toLowerCase();
    const fromDate = $('dateFrom').value || '';
    const toDate = $('dateTo').value || '';

    const inDateRange = (r) => {
        const d = r.Date || '';
        if (fromDate && d < fromDate) return false;
        if (toDate && d > toDate) return false;
        return true;
    };

    if (activeTab === 'blade') {

        $('csvLink').href = '/download/blade_csv';

        $('thead').innerHTML = `
        <tr>
            <th>ID</th>
            <th>Date / Time</th>
            <th>Count</th>
            <th>Item</th>
            <th>Breakdown</th>
        </tr>`;

        const rows = bladeLogs.filter(r =>
            JSON.stringify(r).toLowerCase().includes(q) &&
            inDateRange(r)
        );

        $('tbody').innerHTML = rows.length
            ? rows.map(r => `
            <tr>
                <td><span class="id-badge">${r.ID || ''}</span></td>
                <td style="font-size:10px;color:var(--muted)">
                    ${r.Date || ''}<br>${r.Time || ''}
                </td>
                <td><span class="count-bold">${r["Blade Count"] || 0}</span></td>
                <td style="font-weight:600">${r["Current Item"] || ''}</td>
                <td style="font-size:10px;color:var(--muted)">
                    ${r["Item Breakdown"] || ''}
                </td>
            </tr>`).join('')
            : `
            <tr>
                <td colspan="5" style="text-align:center;color:var(--muted);padding:20px">
                    No blade reset logs yet
                </td>
            </tr>`;

    } else {

        $('csvLink').href = '/download/item_csv';

        $('thead').innerHTML = `
        <tr>
            <th>ID</th>
            <th>Date / Time</th>
            ${ITEM_COLUMNS.map(c => `<th>${c}</th>`).join('')}
        </tr>`;

        const rows = itemLogs.filter(r =>
            JSON.stringify(r).toLowerCase().includes(q) &&
            inDateRange(r)
        );

        $('tbody').innerHTML = rows.length
            ? rows.map(r => `
            <tr>
                <td><span class="id-badge">${r.ID || ''}</span></td>
                <td style="font-size:10px;color:var(--muted)">
                    ${r.Date || ''}<br>${r.Time || ''}
                </td>
                ${ITEM_COLUMNS.map(col => {
                    const value = parseInt(r[col] || 0, 10);
                    return `
                    <td>
                        <span class="item-cell ${value > 0 ? 'nonzero' : ''}">
                            ${value}
                        </span>
                    </td>`;
                }).join('')}
            </tr>`).join('')
            : `
            <tr>
                <td colspan="${ITEM_COLUMNS.length + 2}"
                    style="text-align:center;color:var(--muted);padding:20px">
                    No item wise logs yet
                </td>
            </tr>`;
    }
}

/* Tabs */
$('bladeTab').onclick = () => {
  activeTab = 'blade';
  $('bladeTab').classList.add('active');
  $('itemTab').classList.remove('active');
  renderLogs();
};
$('itemTab').onclick = () => {
  activeTab = 'item';
  $('itemTab').classList.add('active');
  $('bladeTab').classList.remove('active');
  renderLogs();
};
$('search').oninput   = renderLogs;
$('refreshBtn').onclick = loadLogs;
$('dateFrom').onchange  = renderLogs;
$('dateTo').onchange    = renderLogs;
$('clearDateBtn').onclick = () => {
  $('dateFrom').value = '';
  $('dateTo').value   = '';
  renderLogs();
};

/* ── Change Item (no reset) ── */
const changeItemModal = $('changeItemModal');
$('changeItemBtn').onclick = () => {
  $('changeItemSelect').value = state.current_item || items[0];
  changeItemModal.classList.add('open');
};
$('cancelChangeItemBtn').onclick = () => changeItemModal.classList.remove('open');
changeItemModal.addEventListener('click', e => {
  if (e.target === changeItemModal) changeItemModal.classList.remove('open');
});
$('confirmChangeItemBtn').onclick = async () => {
  const new_item = $('changeItemSelect').value;
  if (new_item === state.current_item) {
    changeItemModal.classList.remove('open');
    toast('Already running ' + new_item, '');
    return;
  }
  changeItemModal.classList.remove('open');
  const r = await fetch('/change_item', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ new_item }),
  });
  const d = await r.json();
  if (d.status !== 'ok') return toast(d.message || 'Change failed', 'error');
  await loadLogs();
  toast(`Item switched to: ${new_item}  (blade count unchanged)`, 'info');
};

/* ── Reset Item ── */
$('resetItemBtn').onclick = async () => {
  const next_item = $('itemSelect').value;
  if (!confirm(`Save report for current item and reset its count?\nNext item will be: ${next_item}`)) return;
  const r = await fetch('/reset_item', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ next_item }),
  });
  const d = await r.json();
  if (d.status !== 'ok') return toast(d.message || 'Reset failed', 'error');
  state = d.state;
  renderState();
  await loadLogs();
  toast('Item report saved. Running: ' + next_item, 'success');
};

/* ── Blade Reset Modal ── */
const bladeModal = $('bladeResetModal');
$('resetBladeBtn').onclick = () => {
  $('bladeResetItemSelect').value = state.current_item || items[0];
  bladeModal.classList.add('open');
};
$('cancelBladeResetBtn').onclick = () => bladeModal.classList.remove('open');
bladeModal.addEventListener('click', e => { if (e.target === bladeModal) bladeModal.classList.remove('open'); });
$('confirmBladeResetBtn').onclick = async () => {
  const next_item = $('bladeResetItemSelect').value;
  bladeModal.classList.remove('open');
  const r = await fetch('/reset_blade', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ next_item }),
  });
  const d = await r.json();
  if (d.status !== 'ok') return toast(d.message || 'Blade reset failed', 'error');
  await loadLogs();
  toast(`Blade reset. Now running: ${d.next_item}`, 'success');
};

/* ── SSE ── */
function connectSSE() {
  const es = new EventSource('/events');

  es.addEventListener('metrics', e => {
    state = JSON.parse(e.data);
    renderState();
  });

  es.addEventListener('cut', e => {
    state = JSON.parse(e.data);
    $('lastCut').textContent = new Date().toTimeString().slice(0, 8);
    const el = $('bladeCount');
    el.classList.remove('flash');
    void el.offsetWidth;
    el.classList.add('flash');
    renderState();
  });

  es.addEventListener('blade_reset', e => {
    state = JSON.parse(e.data);
    populateSelects(items, state.current_item);
    renderState();
    loadLogs();
  });

  es.addEventListener('item_reset', e => {
    state = JSON.parse(e.data);
    renderState();
    loadLogs();
  });

  /* item_changed fires when /change_item succeeds (no reset) */
  es.addEventListener('item_changed', e => {
    state = JSON.parse(e.data);
    $('itemSelect').value = state.current_item || items[0];
    renderState();
    loadLogs();
  });

  es.onerror = () => { es.close(); setTimeout(connectSSE, 3000); };
}

/* Boot */
loadState();
loadLogs();
connectSSE();
</script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(DASHBOARD_HTML, rtsp_url=RTSP_URL)

@app.route("/vision")
def vision():
    return render_template_string(VISION_HTML)

if __name__ == "__main__":
    threading.Thread(target=detection_loop,      daemon=True, name="CutCounter").start()
    threading.Thread(target=gstreamer_loop,      daemon=True, name="GStreamer").start()
    threading.Thread(target=billet_refresh_loop, daemon=True, name="BilletDB").start()

    print("=" * 64)
    print("MSP Blade + Item Wise Cut Counter  v3  (Hotsow 1)")
    print("Main Dashboard : http://localhost:5024/")
    print("Vision Feed    : http://localhost:5024/vision")
    print("Blade CSV      : http://localhost:5024/download/blade_csv")
    print("Item CSV       : http://localhost:5024/download/item_csv")
    print("Combined Excel : http://localhost:5024/download/combined_excel")
    print("Live Blade CSV : http://localhost:5024/download/blade_live_csv")
    print("Live Item CSV  : http://localhost:5024/download/item_live_csv")
    print("=" * 64)
    app.run(host="0.0.0.0", port=5024, debug=False, use_reloader=False, threaded=True)